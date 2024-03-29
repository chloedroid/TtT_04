#######################################################################################################################
# A. Importeer csv (uit adlib) naar dataframe (pandas)

# pip install pandas
import pandas as pd

# inlezen csv
# ingeven padnaam naar csv, en indien van toepassing aanvullen delimiter
lijst = pd.read_csv("export (13).csv", delimiter=';')
print(lijst)

#######################################################################################################################
# B. Maak lijsten met ontbrekende velden
# 1. Selecteer ontbrekende velden (via pandas)
# 1.1 Objectnaam

# identificeer wanneer objectnaam 'leeg' is
ontbrekende_objectnaam = pd.isna(lijst['objectnaam'])
print(ontbrekende_objectnaam)

# hou vanuit de lijst enkel de records over waarbij objectnaam leeg = true was
ontbrekende_objectnaam = lijst[ontbrekende_objectnaam]
print(ontbrekende_objectnaam)

# 1.2 Titel

ontbrekende_titel = pd.isna(lijst['titel'])
ontbrekende_titel = lijst[ontbrekende_titel]
print(ontbrekende_titel)

# 1.3 Beschrijving

ontbrekende_beschrijving = pd.isna(lijst['beschrijving'])
ontbrekende_beschrijving = lijst[ontbrekende_beschrijving]
print(ontbrekende_beschrijving)

# 1.4 Vervaardiger

ontbrekende_vervaardiger = pd.isna(lijst['vervaardiger'])
ontbrekende_vervaardiger = lijst[ontbrekende_vervaardiger]
print(ontbrekende_vervaardiger)

# 1.5 Datering

ontbrekende_datering = pd.isna(lijst['vervaardiging.datum.begin'])
ontbrekende_datering = lijst[ontbrekende_datering]
print(ontbrekende_datering)

# 1.6 Vervaardiging rol

ontbrekende_rol = pd.isna(lijst['vervaardiger.rol'])
ontbrekende_rol = lijst[ontbrekende_rol]
print(ontbrekende_rol)

# 1.7 vervaardiging plaats

ontbrekende_plaats = pd.isna(lijst['vervaardiging.plaats'])
ontbrekende_plaats = lijst[ontbrekende_plaats]
print(ontbrekende_plaats)

#######################################################################################################################
# B. Maak lijsten met ontbrekende velden
# 2. Output (excel) maken met lijsten met ontbrekende velden (via openpyxl)

# pip install openpyxl
from openpyxl import Workbook

# maak een excel aan
wb = Workbook()

# voeg een sheet toe met ontbrekende velden
# 2.1 objectnaam

from openpyxl.utils.dataframe import dataframe_to_rows

# maak sheet (tabblad)
ws = wb.create_sheet("Objectnaam")
# zet dataframe (pandas) om naar rijen in het tabblad
rows = dataframe_to_rows(ontbrekende_objectnaam, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# sla excel op
wb.save(r"C:\Users\gelderch\PycharmProjects\TtT_04\output.xlsx")

# 2.2 titel

ws = wb.create_sheet("Titel")
rows = dataframe_to_rows(ontbrekende_titel, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# 2.3 beschrijving

ws = wb.create_sheet("Beschrijving")
rows = dataframe_to_rows(ontbrekende_beschrijving, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# 2.4 vervaardiger

ws = wb.create_sheet("Vervaardiger")
rows = dataframe_to_rows(ontbrekende_vervaardiger, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# 2.5 datering

ws = wb.create_sheet("Datering")
rows = dataframe_to_rows(ontbrekende_datering, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# 2.6 Vervaardiging rol

ws = wb.create_sheet("Rol vervaardiger")
rows = dataframe_to_rows(ontbrekende_rol, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# 2.7 Vervaardiging plaats

ws = wb.create_sheet("Vervaardiging plaats")
rows = dataframe_to_rows(ontbrekende_plaats, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"C:\Users\gelderch\PycharmProjects\TtT_04\output.xlsx")

#######################################################################################################################
# C. visualisatie van de data
# 1. tellen van de data (pandas)

# tel aantal keer objectnummer aanwezig is in lijst records zonder objectnamen (gezien objectnummer altijd aanwezig)
ontbrekende_objectnamen = ontbrekende_objectnaam['objectnummer'].count()
print(ontbrekende_objectnamen)
# doe hetzelfde voor de overige velden
ontbrekende_titels = ontbrekende_titel['objectnummer'].count()
print(ontbrekende_titels)
ontbrekende_beschrijvingen = ontbrekende_beschrijving['objectnummer'].count()
ontbrekende_vervaardigers = ontbrekende_vervaardiger['objectnummer'].count()
ontbrekende_dateringen = ontbrekende_datering['objectnummer'].count()
ontbrekende_rol = ontbrekende_rol['objectnummer'].count()
ontbrekende_plaats = ontbrekende_plaats['objectnummer'].count()

#######################################################################################################################
# C. visualisatie van de data
# 2. weergeven van de data in grafiek in excel (openpyxl)

# 2.1 voeg de data toe in excel
# zet de data in een list
labels = ["objectnaam", "titel", "beschrijving", "vervaardiger", "datering", "rol", "plaats"]
ontbrekende_data = [ontbrekende_objectnamen, ontbrekende_titels, ontbrekende_beschrijvingen, ontbrekende_vervaardigers,
                    ontbrekende_dateringen, ontbrekende_rol, ontbrekende_plaats]

# voeg de list toe aan de excel (openpyxl)

ws = wb.active
ws.title = 'Info'
# titel lijst
ws['A1'] = "Ontbrekende velden"

# de labels
rij1 = 2
for label in labels:
    ws.cell(row=rij1, column=1).value = label
    rij1 += 1

# de waardes
rij2 = 2
for veld in ontbrekende_data:
    ws.cell(row=rij2, column=2).value = veld
    rij2 += 1

# 2.2 maak de grafiek in excel

from openpyxl.chart import BarChart3D, Reference

# selecteer de data
data = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=8)

# selecteer de labels
labels = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=8)

# creeer grafiek
chart = BarChart3D()

# voeg grafiek toe aan excel
ws.add_chart(chart, "E2")

# voeg titel toe
chart.title = 'Ontbrekende data'

# voeg namen x en y as toe
chart.y_axis.title = 'aantal'
chart.x_axis.title = 'velden'

# voeg de data & labels toe
chart.add_data(data)
chart.set_categories(labels)

# sla excel op
wb.save(r"C:\Users\gelderch\PycharmProjects\TtT_04\output.xlsx")